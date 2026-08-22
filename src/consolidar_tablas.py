

import colorsys
import math
import os
import re
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIGURACION
# =============================================================================

RAIZ = Path(__file__).resolve().parent.parent
CARPETA_ENTRADA = RAIZ / "input" / "tablas" / "input"
CARPETA_SALIDA = RAIZ / "input" / "tablas" / "output"

HOJAS_NC = [
    'NO_CONVENCIONAL_10C',
    'NO_CONVENCIONAL_7C',
]

HOJAS = [
    'T1_RESIDENCIAL_7C',
    'T1_RESIDENCIAL_10C_COND_0',
    'T1_RESIDENCIAL_10C_COND_9',
    'T2_EDIFICIOS_7C',
    'T2_EDIFICIOS_10C',
    'T3_COMERCIAL_7C',
    'T3_COMERCIAL_10C',
]

# Nombres posibles de cada columna (en minuscula, sin acentos) - hojas NC
ALIAS = {
    'DESTANEX': ['codigo', 'destanex', 'cod'],
    'USO_LADM': ['uso ladm', 'uso_ladm', 'uso'],
    'TIPO':     ['tipo'],
    'VALOR':    ['valor adoptado', 'valor_adoptado', 'adoptado', 'valor'],
}

FILA_SUFIJO = 5   # fila Excel 6 -> indice 5  (rotulo del bloque)
FILA_HEADER = 6   # fila Excel 7 -> indice 6  ("Puntaje" | "Adoptados")
FILA_DATOS = 7    # fila Excel 8 -> indice 7  (primer dato)
N_PUNTAJES = 100

# El sufijo NO se lee del rotulo: se asigna por posicion segun el tipo de hoja.
SUFIJOS_POR_TIPO = {
    'RESIDENCIAL': ['011', '012', '013', '014', '015', '016'],
    'EDIFICIOS':   ['011', '012', '013', '014', '015', '016'],
    'COMERCIAL':   ['021', '022', '023'],
    'INDUSTRIAL':  ['031', '032', '033'],
}


def tabla_valor_vigente():
    """El Tablas_Valor_Consolidado_V1_<fecha>.xlsx mas reciente, o None."""
    if not CARPETA_SALIDA.is_dir():
        return None
    # Se descartan los "~$": son los bloqueos que deja Excel al abrir un archivo.
    hallados = sorted((f for f in CARPETA_SALIDA.glob("Tablas_Valor_Consolidado_*.xlsx")
                       if not f.name.startswith("~$")),
                      key=lambda f: f.stat().st_mtime, reverse=True)
    return hallados[0] if hallados else None


def consolidado_mas_reciente():
    """El Consolidado_<fecha>.xlsx mas reciente de la carpeta de entrada."""
    if not CARPETA_ENTRADA.is_dir():
        return None
    hallados = sorted((f for f in CARPETA_ENTRADA.glob("Consolidado_*.xlsx")
                       if not f.name.startswith("~$")),
                      key=lambda f: f.stat().st_mtime, reverse=True)
    return hallados[0] if hallados else None


# =============================================================================
# UTILIDADES
# =============================================================================

def norm(t) -> str:
    """Normaliza un texto: minuscula, sin acentos, sin dobles espacios."""
    t = str(t).strip().lower()
    for a, b in zip('aeioun', 'aeioun'):
        t = t.replace(a, b)
    for a, b in zip('\xe1\xe9\xed\xf3\xfa\xf1', 'aeioun'):
        t = t.replace(a, b)
    return re.sub(r'\s+', ' ', t)


VACIOS = ('', 'nan', 'none', 'nat', '-', '#n/a', '#\xa1valor!', '#value!', '#ref!')


def a_float(v):
    """
    Convierte la celda a float SIN perder ni inventar digitos. None si no aplica.
    """
    if v is None or isinstance(v, bool):
        return None

    # Si la celda YA es numero no se toca su escritura. Aqui estaba el bug: a un
    # float como 1185266.6590197692 se le borraba el punto creyendo que era
    # separador de miles -> 11852666590197692.
    if isinstance(v, (int, float, Decimal)):
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f

    texto = str(v).replace('$', '').replace('\xa0', '').replace(' ', '').strip()
    if texto.lower() in VACIOS:
        return None

    negativo = texto.startswith('(') and texto.endswith(')')   # formato contable
    texto = texto.strip('()')

    hay_punto, hay_coma = '.' in texto, ',' in texto
    if hay_punto and hay_coma:
        # el separador decimal es el que aparece de ultimo
        dec = ',' if texto.rfind(',') > texto.rfind('.') else '.'
        mil = '.' if dec == ',' else ','
        texto = texto.replace(mil, '').replace(dec, '.')
    elif hay_coma:
        ent, _, frac = texto.rpartition(',')
        texto = texto.replace(',', '') if len(frac) == 3 else f'{ent}.{frac}'
    elif hay_punto:
        _, _, frac = texto.rpartition('.')
        texto = texto.replace('.', '') if len(frac) == 3 else texto

    try:
        f = float(texto)
    except (ValueError, TypeError):
        return None
    if math.isnan(f) or math.isinf(f):
        return None
    return -f if negativo else f


def limpiar_valor(v) -> int:
    """'$ 1.234.567,00' -> 1234567. Redondeo comercial; 0 si no se puede."""
    f = a_float(v)
    if f is None:
        return 0
    return int(Decimal(str(f)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))


# =============================================================================
# CONTROL DE VALORES
# =============================================================================

VALOR_MIN = 10_000        # piso razonable de $/m2
VALOR_MAX = 50_000_000    # techo razonable de $/m2
SALTO_MAX = 2.0           # un puntaje no deberia valer mas del doble del anterior

GRAVES = ('FUERA_DE_RANGO', 'NO_CONVERTIBLE', 'VALOR_CERO', 'NO_CRECIENTE')

REPORTE = []              # lo llena auditar_valores() en cada tabla
SIN_PALETA = set()        # familias sin color asignado


def _es_texto(v):
    return isinstance(v, str) and str(v).strip().lower() not in VACIOS


def auditar_valores(nombre, crudos, valores, etiquetas=None, serie=True) -> list:
    """Compara la celda como venia contra el valor limpio y acumula alertas."""
    crudos, valores = list(crudos), list(valores)
    etiquetas = list(etiquetas) if etiquetas is not None else list(
        range(1, len(valores) + 1))

    def anota(fila, alerta, detalle, crudo=None, valor=None):
        REPORTE.append({'TABLA': nombre, 'FILA': fila, 'ALERTA': alerta,
                        'DETALLE': detalle,
                        'VALOR_ORIGEN': '' if crudo is None else repr(crudo),
                        'VALOR_FINAL': valor})

    # Como viene escrita la MAYORIA de la tabla. Lo que se salga de ese formato
    # es justamente lo que hay que revisar.
    n_txt = sum(1 for c in crudos if _es_texto(c))
    n_num = sum(1 for c in crudos if isinstance(c, (int, float, Decimal))
                and not isinstance(c, bool) and a_float(c) is not None)
    formato = 'texto' if n_txt > n_num else 'numero'

    for i, (crudo, valor) in enumerate(zip(crudos, valores)):
        fila = etiquetas[i]
        f = a_float(crudo)

        if f is None:
            if _es_texto(crudo):
                anota(fila, 'NO_CONVERTIBLE',
                      f'no se pudo leer como numero: {crudo!r}', crudo, valor)
            else:
                anota(fila, 'VALOR_CERO', 'celda vacia', crudo, valor)
            continue

        propio = 'texto' if _es_texto(crudo) else 'numero'
        if propio != formato:
            anota(fila, 'FORMATO_DISTINTO',
                  f'viene como {propio} y el resto de la tabla como {formato}',
                  crudo, valor)

        # venia con decimales -> se redondeo (aqui caia T3_COMERCIAL_10C_021)
        if not float(f).is_integer():
            anota(fila, 'DECIMAL_REDONDEADO', f'{f!r} -> {valor:,}', crudo, valor)

        if valor == 0:
            anota(fila, 'VALOR_CERO', 'quedo en cero', crudo, valor)
        elif not (VALOR_MIN <= valor <= VALOR_MAX):
            anota(fila, 'FUERA_DE_RANGO',
                  f'{valor:,} fuera de [{VALOR_MIN:,} - {VALOR_MAX:,}]', crudo, valor)

    # la serie de puntajes debe crecer y sin saltos de magnitud
    if serie:
        for i, (a, b) in enumerate(zip(valores, valores[1:])):
            if not a or not b:
                continue
            if b < a:
                anota(etiquetas[i + 1], 'NO_CRECIENTE', f'{a:,} -> {b:,}',
                      crudos[i + 1], b)
            elif b > SALTO_MAX * a:
                anota(etiquetas[i + 1], 'SALTO_ATIPICO',
                      f'{a:,} -> {b:,}  (x{b / a:,.1f})', crudos[i + 1], b)

    return valores


def resumen_control() -> pd.DataFrame:
    """Imprime el resumen por tabla y devuelve el DataFrame del reporte."""
    df = pd.DataFrame(REPORTE, columns=['TABLA', 'FILA', 'ALERTA', 'DETALLE',
                                        'VALOR_ORIGEN', 'VALOR_FINAL'])
    print("=" * 60)
    print("  CONTROL DE VALORES")
    print("=" * 60)
    if df.empty:
        print("OK: sin alertas, todas las tablas vienen igual.")
    else:
        for (tabla, alerta), n in df.groupby(['TABLA', 'ALERTA']).size().items():
            icono = '[GRAVE]' if alerta in GRAVES else '[aviso]'
            print(f"  {icono} {tabla:32s} {alerta:20s} {n:>4} caso(s)")
        print(f"\n  Total: {len(df)} alertas en {df['TABLA'].nunique()} tabla(s). "
              f"Detalle completo en la hoja CONTROL.")
    return df


# =============================================================================
# HOJAS NO CONVENCIONALES
# =============================================================================

def ubicar_encabezado(df_raw):
    """(indice_fila, {campo: indice_columna}) de la fila con 'Valor Adoptado'."""
    for i in range(min(30, df_raw.shape[0])):
        celdas = [norm(x) for x in df_raw.iloc[i].tolist()]
        if any(c in ALIAS['VALOR'] for c in celdas):
            cols = {}
            for campo, opciones in ALIAS.items():
                for c, texto in enumerate(celdas):
                    if texto in opciones:
                        cols[campo] = c
                        break
            return i, cols
    return None, {}


def procesar_hoja_nc(xls, nombre_hoja: str) -> pd.DataFrame:
    df_raw = pd.read_excel(xls, sheet_name=nombre_hoja, header=None)
    f_hdr, cols = ubicar_encabezado(df_raw)

    if f_hdr is None:
        print(f"  AVISO: no se encontro el encabezado en '{nombre_hoja}'")
        return pd.DataFrame()

    faltan = [k for k in ('DESTANEX', 'TIPO', 'VALOR') if k not in cols]
    if faltan:
        print(f"  AVISO: '{nombre_hoja}' sin columnas {faltan}")
        return pd.DataFrame()

    print(f"  Encabezado en fila Excel {f_hdr + 1} -> {cols}")

    filas, crudos_nc = [], []
    for i in range(f_hdr + 1, df_raw.shape[0]):
        cod = df_raw.iat[i, cols['DESTANEX']]
        tip = df_raw.iat[i, cols['TIPO']]
        val = df_raw.iat[i, cols['VALOR']]

        if pd.isna(cod) and pd.isna(tip) and pd.isna(val):
            continue

        d_cod = re.sub(r'\D', '', str(cod))
        d_tip = re.sub(r'\D', '', str(tip))
        if not d_cod or not d_tip:
            continue

        crudos_nc.append(val)
        filas.append({
            'HOJA':     nombre_hoja,
            'DESTANEX': d_cod.zfill(3),
            'USO_LADM': str(df_raw.iat[i, cols['USO_LADM']]).strip()
                        if 'USO_LADM' in cols else '',
            'TIPO':     int(d_tip),
            'VALOR':    limpiar_valor(val),
        })

    df = pd.DataFrame(filas)
    if df.empty:
        print(f"  AVISO: '{nombre_hoja}' con 0 filas leidas")
        return df

    auditar_valores(nombre_hoja, crudos_nc, df['VALOR'].tolist(),
                    etiquetas=(df['DESTANEX'] + '_' + df['TIPO'].astype(str)).tolist(),
                    serie=False)

    ceros = (df['VALOR'] == 0).sum()
    if ceros:
        print(f"  AVISO: {ceros} filas con valor 0")

    dups = df.duplicated(subset=['DESTANEX', 'TIPO'], keep=False)
    if dups.any():
        print(f"  AVISO: {dups.sum()} filas con DESTANEX+TIPO repetido:")
        print(df[dups][['DESTANEX', 'USO_LADM', 'TIPO', 'VALOR']].to_string(index=False))

    print(f"  OK {nombre_hoja:25s} -> {len(df)} filas, "
          f"{df['DESTANEX'].nunique()} codigos, "
          f"[{df['VALOR'].min():,} ... {df['VALOR'].max():,}]")
    return df


def leer_no_convencionales(xls) -> pd.DataFrame:
    print("=" * 60)
    print("  TABLAS NO CONVENCIONALES")
    print("=" * 60)

    partes = []
    for nombre_hoja in HOJAS_NC:
        if nombre_hoja not in xls.sheet_names:
            print(f"AVISO: hoja '{nombre_hoja}' no encontrada, se salta.\n")
            continue
        print(f"Hoja: '{nombre_hoja}'")
        parte = procesar_hoja_nc(xls, nombre_hoja)
        if not parte.empty:
            partes.append(parte)
        print()

    columnas_vacias = ['HOJA', 'DESTANEX', 'USO_LADM', 'TIPO', 'VALOR', 'CLAVE']
    if not partes:
        print("AVISO: ninguna hoja NC produjo datos.\n")
        return pd.DataFrame(columns=columnas_vacias)

    df_nc = pd.concat(partes, ignore_index=True)
    df_nc['CLAVE'] = df_nc['DESTANEX'] + '_' + df_nc['TIPO'].astype(str)
    print(f"OK df_nc: {df_nc.shape[0]} filas x {df_nc.shape[1]} columnas\n")
    return df_nc


# =============================================================================
# HOJAS CONVENCIONALES
# =============================================================================

def sufijos_de_hoja(nombre_hoja: str) -> list:
    """
    Sufijos que le corresponden a la hoja segun su nombre. Compara por token
    exacto para que 'T11_CCOMERCIALES_7C' no caiga en la regla de 'COMERCIAL'.
    """
    tokens = nombre_hoja.upper().split('_')
    for palabra, sufijos in SUFIJOS_POR_TIPO.items():
        if palabra in tokens:
            return sufijos
    return []   # sin sufijo -> una sola tabla con el nombre de la hoja


def ubicar_filas(df_raw):
    """Ubica las filas buscando 'Puntaje'; si no aparece, usa las constantes."""
    for i in range(df_raw.shape[0]):
        fila = df_raw.iloc[i].astype(str).str.strip().str.lower()
        if fila.eq('puntaje').any():
            return i - 1, i, i + 1
    return FILA_SUFIJO, FILA_HEADER, FILA_DATOS


def procesar_hoja(xls, nombre_hoja: str) -> dict:
    tablas = {}
    df_raw = pd.read_excel(xls, sheet_name=nombre_hoja, header=None)
    f_suf, f_hdr, f_dat = ubicar_filas(df_raw)

    # Los bloques van separados por una columna vacia: 1, 4, 7, 10 ... y los
    # adoptados quedan siempre en la columna siguiente (c + 1).
    cols_puntaje = [c for c in range(df_raw.shape[1])
                    if str(df_raw.iat[f_hdr, c]).strip().lower() == 'puntaje'
                    and c + 1 < df_raw.shape[1]]

    if not cols_puntaje:
        print(f"  AVISO: sin pares Puntaje|Adoptados en '{nombre_hoja}'")
        return tablas

    sufijos = sufijos_de_hoja(nombre_hoja)
    if sufijos:
        if len(cols_puntaje) != len(sufijos):
            print(f"  AVISO: '{nombre_hoja}' esperaba {len(sufijos)} bloques y hay "
                  f"{len(cols_puntaje)}. Se asignan los primeros "
                  f"{min(len(sufijos), len(cols_puntaje))}.")
        nombres = [f"{nombre_hoja}_{s}" for s in sufijos]
    else:
        if len(cols_puntaje) > 1:
            print(f"  AVISO: '{nombre_hoja}' sin sufijos pero con "
                  f"{len(cols_puntaje)} bloques. Se usa solo el primero.")
        cols_puntaje = cols_puntaje[:1]
        nombres = [nombre_hoja]

    for nombre, c in zip(nombres, cols_puntaje):
        # El rotulo de la fila 6 deberia terminar en el sufijo asignado.
        rotulo = str(df_raw.iat[f_suf, c])
        digitos = re.findall(r'\d{3}', rotulo)
        if digitos and not nombre.endswith(digitos[-1]):
            print(f"  AVISO: {nombre} tiene rotulo '{rotulo.strip()}' "
                  f"(sufijo {digitos[-1]}) y no coincide con la posicion.")

        puntajes = df_raw.iloc[f_dat:f_dat + N_PUNTAJES, c].tolist()
        crudos = df_raw.iloc[f_dat:f_dat + N_PUNTAJES, c + 1].tolist()
        valores = [limpiar_valor(v) for v in crudos]

        if puntajes != list(range(1, N_PUNTAJES + 1)):
            print(f"  AVISO: {nombre} con puntajes que no van 1..{N_PUNTAJES}")
        if len(valores) < N_PUNTAJES:
            print(f"  AVISO: {nombre} con solo {len(valores)} valores, se rellena")
            crudos += [None] * (N_PUNTAJES - len(crudos))
            valores += [0] * (N_PUNTAJES - len(valores))
        if 0 in valores:
            print(f"  AVISO: {nombre} con {valores.count(0)} valores en 0")

        auditar_valores(nombre, crudos, valores)
        tablas[nombre] = valores
        print(f"  OK {nombre:30s} ->  [{valores[0]:,} ... {valores[-1]:,}]  "
              f"({len(valores)} valores)")

    return tablas


def leer_convencionales(xls) -> pd.DataFrame:
    print("=" * 60)
    print("  TABLAS CONVENCIONALES")
    print("=" * 60)

    todas = {}
    for nombre_hoja in HOJAS:
        if nombre_hoja not in xls.sheet_names:
            print(f"AVISO: hoja '{nombre_hoja}' no encontrada, se salta.\n")
            continue
        print(f"Hoja: '{nombre_hoja}'")
        todas.update(procesar_hoja(xls, nombre_hoja))
        print()

    df_tablas = pd.DataFrame({'PUNTAJE': range(1, N_PUNTAJES + 1)})
    for nombre_col, valores in todas.items():
        vals = list(valores)[:N_PUNTAJES]
        df_tablas[nombre_col] = vals + [0] * (N_PUNTAJES - len(vals))

    print(f"OK: {df_tablas.shape[0]} filas x {df_tablas.shape[1]} columnas "
          f"({df_tablas.shape[1] - 1} tablas)\n")
    return df_tablas


# =============================================================================
# PALETA
# =============================================================================
# Color BASE por familia. El cuerpo y el encabezado se derivan aclarandolo, y
# el estrato define que tan claro queda, para distinguir los bloques.

FAMILIAS = {
    'T1_RESIDENCIAL_7C':          '2E86C1',   # azul
    'T1_RESIDENCIAL_10C_COND_0':  '27AE60',   # verde
    'T1_RESIDENCIAL_10C_COND_9':  '16A085',   # verde azulado
    'T1_RESIDENCIAL_10C_0':       '27AE60',
    'T1_RESIDENCIAL_10C_9':       '16A085',
    'T2_EDIFICIOS_7C':            'E67E22',   # naranja
    'T2_EDIFICIOS_10C':           'CA6F1E',
    'T3_COMERCIAL_7C':            'C0392B',   # rojo
    'T3_COMERCIAL_10C':           'E74C3C',
    'T4_INDUSTRIAL_7C':           '7D3C98',   # violeta
    'T4_INDUSTRIAL_10C':          '9B59B6',
    'T5_INSTITUCIONAL_ED_7C':     'B7950B',   # mostaza
    'T6_INSTITUCIONAL_SA_7C':     '148F77',   # esmeralda
    'T7_INSTITUCIONAL_SER_7C':    '5D6D7E',   # gris azulado
    'T8_INSTITUCIONAL_IG_7C':     '76448A',
    'T9_HOTELES_7C':              'AD1457',   # rosa fuerte
    'T11_CCOMERCIALES_7C':        '922B21',
    'T13_UNIDAD_DEPORTIVA_7C':    '1D8348',   # verde bosque
}

# Fraccion de blanco mezclada: indice 0 = estrato 1 (mas claro) ... 5 = estrato 6
F_BODY = [0.90, 0.86, 0.82, 0.78, 0.74, 0.70]   # celdas de datos
F_HDR = [0.62, 0.56, 0.50, 0.44, 0.38, 0.32]    # encabezado


def _hex_rgb(h: str):
    return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))


def _mezclar(base_hex: str, f: float, hacia: str = 'FFFFFF') -> str:
    """Mezcla base_hex con 'hacia' en proporcion f (0 = base puro)."""
    a, b = _hex_rgb(base_hex), _hex_rgb(hacia)
    return '{:02X}{:02X}{:02X}'.format(
        *(int(round(a[i] + (b[i] - a[i]) * f)) for i in range(3)))


def _base_automatico(familia: str) -> str:
    """Tono estable derivado del nombre, para familias fuera de FAMILIAS."""
    h = (sum(ord(c) * (i + 7) for i, c in enumerate(familia)) % 360) / 360
    r, g, b = colorsys.hls_to_rgb(h, 0.42, 0.55)
    return '{:02X}{:02X}{:02X}'.format(int(r * 255), int(g * 255), int(b * 255))


def _texto_contraste(fondo_hex: str) -> str:
    """Negro o blanco segun la luminancia del fondo."""
    r, g, b = _hex_rgb(fondo_hex)
    lum = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    return '1A1A1A' if lum > 0.55 else 'FFFFFF'


def get_color(col_name: str):
    """(color_body, color_header) de una columna como 'T1_RESIDENCIAL_7C_011'."""
    nombre = col_name.upper()
    m = re.search(r'_(\d{3})$', nombre)
    sufijo = m.group(1) if m else None
    familia = nombre[:m.start()] if m else nombre

    # Prefijo mas especifico primero, para que '_10C_COND_0' gane sobre '_10C'
    base = None
    for prefijo in sorted(FAMILIAS, key=len, reverse=True):
        if familia.startswith(prefijo):
            base = FAMILIAS[prefijo]
            break
    if base is None:
        SIN_PALETA.add(familia)
        base = _base_automatico(familia)

    # Paso del degradado: ultimo digito del sufijo (011->0, 016->5, 021->0)
    paso = (int(sufijo[-1]) - 1) if sufijo else 2
    paso = max(0, min(paso, len(F_BODY) - 1))
    return _mezclar(base, F_BODY[paso]), _mezclar(base, F_HDR[paso])


COLOR_PUNTAJE_HDR = '2C3E50'
COLOR_PUNTAJE_BODY = 'EBF5FB'
COLOR_CLAVE_HDR = '2C3E50'
COLOR_CLAVE_BODY = 'EBF5FB'
COLOR_NC_HDR = 'C39BD3'
COLOR_NC_BODY = 'F5EEF8'
COLOR_CTRL_HDR = '2C3E50'
COLOR_CTRL_GRAVE = 'FADBD8'   # hay que corregir el consolidado
COLOR_CTRL_LEVE = 'FCF3CF'    # se ajusto solo, queda el registro

ANCHO_MINIMO = 14
ANCHO_PUNTAJE = 10
ANCHO_CLAVE = 16
HOLGURA = 3

HOJA_CONV = "CONVENCIONALES"
HOJA_NC = "NO_CONVENCIONALES"
HOJA_CTRL = "CONTROL"


# =============================================================================
# ESCRITURA
# =============================================================================

def _formatear(ruta, df_tablas, df_nc_out, df_control):
    """Colores, anchos y bordes sobre el libro ya escrito."""
    wb = load_workbook(ruta)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fmt_numero = '#,##0'

    # --- CONVENCIONALES ---
    ws = wb[HOJA_CONV]
    columnas = list(df_tablas.columns)
    for col_idx, col_name in enumerate(columnas, start=1):
        letra = get_column_letter(col_idx)
        es_puntaje = (col_name == 'PUNTAJE')
        if es_puntaje:
            color_hdr, color_body = COLOR_PUNTAJE_HDR, COLOR_PUNTAJE_BODY
            ancho = ANCHO_PUNTAJE
        else:
            color_body, color_hdr = get_color(col_name)
            ancho = max(len(col_name) + HOLGURA, ANCHO_MINIMO)

        cell_hdr = ws[f'{letra}1']
        cell_hdr.fill = PatternFill('solid', fgColor=color_hdr)
        cell_hdr.font = Font(bold=True, color=_texto_contraste(color_hdr), size=10)
        cell_hdr.alignment = Alignment(horizontal='center', vertical='center')
        cell_hdr.border = border

        fill_body = PatternFill('solid', fgColor=color_body)
        for row_idx in range(2, len(df_tablas) + 2):
            cell = ws[f'{letra}{row_idx}']
            cell.fill = fill_body
            cell.border = border
            cell.alignment = Alignment(horizontal='right', vertical='center')
            if not es_puntaje:
                cell.number_format = fmt_numero
        ws.column_dimensions[letra].width = ancho

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'B2'

    # --- NO_CONVENCIONALES ---
    ws_nc = wb[HOJA_NC]
    n_nc = len(df_nc_out)
    for col_idx, col_name in enumerate(df_nc_out.columns, start=1):
        letra = get_column_letter(col_idx)
        es_clave = (col_name == 'CLAVE')
        if es_clave:
            color_hdr, color_body = COLOR_CLAVE_HDR, COLOR_CLAVE_BODY
            alin, fmt, ancho = 'left', None, ANCHO_CLAVE
        else:
            color_hdr, color_body = COLOR_NC_HDR, COLOR_NC_BODY
            alin, fmt = 'right', fmt_numero
            ancho = max(len(str(col_name)) + HOLGURA, ANCHO_MINIMO)

        cell_hdr = ws_nc[f'{letra}1']
        cell_hdr.fill = PatternFill('solid', fgColor=color_hdr)
        cell_hdr.font = Font(bold=True, color=_texto_contraste(color_hdr), size=10)
        cell_hdr.alignment = Alignment(horizontal='center', vertical='center')
        cell_hdr.border = border

        fill_body = PatternFill('solid', fgColor=color_body)
        for row_idx in range(2, n_nc + 2):
            cell = ws_nc[f'{letra}{row_idx}']
            cell.fill = fill_body
            cell.border = border
            cell.alignment = Alignment(horizontal=alin, vertical='center')
            if fmt:
                cell.number_format = fmt
        ws_nc.column_dimensions[letra].width = ancho

    ws_nc.row_dimensions[1].height = 22
    ws_nc.freeze_panes = 'B2'

    # --- CONTROL ---
    ws_ct = wb[HOJA_CTRL]
    n_ct = len(df_control)
    cols_ct = list(df_control.columns)
    anchos_ct = {'TABLA': 32, 'FILA': 10, 'ALERTA': 20, 'DETALLE': 48,
                 'VALOR_ORIGEN': 24, 'VALOR_FINAL': 16}
    for col_idx, col_name in enumerate(cols_ct, start=1):
        letra = get_column_letter(col_idx)
        cell_hdr = ws_ct[f'{letra}1']
        cell_hdr.fill = PatternFill('solid', fgColor=COLOR_CTRL_HDR)
        cell_hdr.font = Font(bold=True, color=_texto_contraste(COLOR_CTRL_HDR),
                             size=10)
        cell_hdr.alignment = Alignment(horizontal='center', vertical='center')
        cell_hdr.border = border

        for row_idx in range(2, n_ct + 2):
            cell = ws_ct[f'{letra}{row_idx}']
            alerta = ws_ct[f'C{row_idx}'].value
            cell.fill = PatternFill('solid',
                                    fgColor=(COLOR_CTRL_GRAVE if alerta in GRAVES
                                             else COLOR_CTRL_LEVE))
            cell.border = border
            cell.alignment = Alignment(
                horizontal='right' if col_name in ('FILA', 'VALOR_FINAL') else 'left',
                vertical='center')
            if col_name == 'VALOR_FINAL':
                cell.number_format = fmt_numero
        ws_ct.column_dimensions[letra].width = anchos_ct.get(col_name, ANCHO_MINIMO)

    ws_ct.row_dimensions[1].height = 22
    ws_ct.freeze_panes = 'A2'
    if n_ct:
        ws_ct.auto_filter.ref = f'A1:{get_column_letter(len(cols_ct))}{n_ct + 1}'

    try:
        wb.save(ruta)
    except PermissionError:
        raise SystemExit(f"No se pudo guardar '{ruta}'. Cierrelo en Excel y "
                         f"vuelva a ejecutar.")


def consolidar_tablas(ruta_entrada=None, ruta_salida=None) -> str:
    """
    Arma el consolidado que lee la liquidacion y devuelve su ruta.

    Sin argumentos toma el Consolidado_<fecha>.xlsx mas reciente de
    input/tablas/input/ y escribe en input/tablas/output/.
    """
    # El reporte y la paleta son estado de modulo: se limpian por corrida para
    # que llamar dos veces no duplique las alertas.
    REPORTE.clear()
    SIN_PALETA.clear()

    entrada = Path(ruta_entrada) if ruta_entrada else consolidado_mas_reciente()
    if entrada is None or not entrada.exists():
        raise SystemExit(
            f"No se encontro ningun Consolidado_<fecha>.xlsx en "
            f"{CARPETA_ENTRADA}. Deje ahi el libro que entrega el equipo de "
            f"tablas y vuelva a ejecutar.")

    print("=" * 60)
    print("  CONSOLIDAR TABLAS DE VALOR")
    print("=" * 60)
    print(f"Entrada: {entrada}")

    xls = pd.ExcelFile(entrada)
    print(f"Hojas disponibles: {xls.sheet_names}\n")

    df_nc = leer_no_convencionales(xls)
    df_tablas = leer_convencionales(xls)

    # --- No convencionales: CLAVE + una columna por tabla ---
    if not df_nc.empty:
        df_nc_out = (df_nc
                     .pivot_table(index='CLAVE', columns='HOJA', values='VALOR',
                                  aggfunc='first')
                     .reset_index())
        df_nc_out = df_nc_out[['CLAVE'] + [h for h in HOJAS_NC
                                           if h in df_nc_out.columns]]
    else:
        df_nc_out = pd.DataFrame(columns=['CLAVE'] + HOJAS_NC)
        print("AVISO: df_nc vacio, la hoja NO_CONVENCIONALES saldra sin datos.")

    df_control = resumen_control()

    if ruta_salida:
        salida = Path(ruta_salida)
    else:
        fecha = datetime.now().strftime("%Y%m%d")
        salida = CARPETA_SALIDA / f"Tablas_Valor_Consolidado_V1_{fecha}.xlsx"
    salida.parent.mkdir(parents=True, exist_ok=True)

    try:
        with pd.ExcelWriter(salida, engine='openpyxl') as writer:
            df_tablas.to_excel(writer, index=False, sheet_name=HOJA_CONV)
            df_nc_out.to_excel(writer, index=False, sheet_name=HOJA_NC)
            df_control.to_excel(writer, index=False, sheet_name=HOJA_CTRL)
    except PermissionError:
        raise SystemExit(f"No se pudo escribir '{salida}'. Cierrelo en Excel y "
                         f"vuelva a ejecutar.")

    print("Aplicando formato...")
    _formatear(salida, df_tablas, df_nc_out, df_control)

    print(f"\nOK: {salida}")
    print(f"   {HOJA_CONV}: {len(df_tablas)} filas x {len(df_tablas.columns)} columnas")
    print(f"   {HOJA_NC}:   {len(df_nc_out)} filas x {len(df_nc_out.columns)} columnas")
    if len(df_control):
        print(f"   {HOJA_CTRL}:      {len(df_control)} alertas en "
              f"{df_control['TABLA'].nunique()} tabla(s) -> revisela antes de liquidar")
    else:
        print(f"   {HOJA_CTRL}:      sin alertas")
    if SIN_PALETA:
        print(f"   AVISO: {len(SIN_PALETA)} familias sin color en FAMILIAS "
              f"(se asigno uno automatico): {sorted(SIN_PALETA)}")
    return str(salida)


if __name__ == "__main__":
    import sys
    consolidar_tablas(sys.argv[1] if len(sys.argv) > 1 else None)
