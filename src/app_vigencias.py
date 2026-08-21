"""
=====================================================================
COMPARACION DE VIGENCIAS - version interactiva (Streamlit)
=====================================================================

Lo mismo que entrega comparacion_vigencia.py en Excel, pero como app: se abre
en el navegador desde cualquier PC con el enlace y los percentiles y los
graficos se recalculan sobre lo que el usuario filtre. En el libro de Excel los
bloques vienen fijos (tabla x actividad economica); aqui se puede pedir, por
ejemplo, el percentil 90 de la comuna 17 en T1_RESIDENCIAL_013, que en el Excel
no existe y habria que sacarlo a mano.

Dos hojas, con los mismos filtros al lado en las dos:

    Tablas          el resumen por grupo -cuantos predios, las dos medianas,
                    la diferencia en pesos y cuantos suben o bajan, con fila de
                    TOTAL- y los percentiles 10-100 del total de la seleccion,
                    tambien con la diferencia y su variacion. Un solo boton
                    baja todo en un Excel, incluidos los percentiles abiertos
                    por grupo, que en pantalla no se muestran para no repetir.
    Graficos        las mismas dos curvas del reporte, dibujadas sobre lo
                    filtrado, mas el reparto de la variacion.

Arriba se eligen dos cosas y toda la pantalla responde a ellas:

    MEDIDA   VALOR POR M2 de la construccion, VALOR TOTAL CONSTRUIDO (ese m2
             por el area) o AVALUO del predio completo
    BASE     CATASTRAL, que es sobre lo que se cobra, o COMERCIAL, que es lo
             que se estima que vale el inmueble

El comercial de la VIGENCIA sale de dividir el catastral por el factor de la
comuna: 0.7 en las actualizadas en 2024-2025 y 0.6 en las demas (el terreno va
siempre por 0.7). El de la LIQUIDACION no necesita conversion: el VM2 de las
tablas ya viene comercial. Por eso la variacion comercial no es igual a la
catastral -la vigencia se convierte con dos factores distintos segun la comuna
y la liquidacion con uno solo-, y por eso vale la pena mirar las dos.

Dos fuentes, y de cual haya depende lo que se ve
-----------------------------------------------
output/COMPARACION_VIGENCIA_PUBLICO.parquet   (6.5 MB, SIN identificadores)
    Es lo que mueve las hojas Tablas, Graficos y Reglas: las mismas filas del
    reporte pero sin ID_PREDIO ni numero predial, solo comuna, tabla, actividad
    y valores. Este si va al repositorio y es lo unico que la app necesita.

output/COMPARACION_VIGENCIA_DETALLE.parquet   (44.7 MB, CON identificadores)
    Es lo que mueve la hoja Detalle: 288 mil construcciones fila a fila, con
    ID_PREDIO, numero predial, area, puntaje y las tres medidas en las dos
    bases. Lo escribe comparacion_vigencia.py al lado del otro.

    OJO: el .gitignore lo deja FUERA a proposito, porque este repositorio es
    publico. Corriendo la app en local el detalle esta y la hoja funciona; en
    el deploy de Streamlit Cloud el archivo no existe y la hoja lo dice en vez
    de fallar. Para tenerlo publicado hay que decidirlo antes: el repositorio
    tendria que volver a ser privado, o la app quedar detras de contrasena.
    Subirlo tal como esta hoy publica el numero predial de 288 mil predios, y
    del historial de git eso no se saca borrando el archivo despues.

Trae solo los predios comparables (una sola construccion y valor de tabla en
las dos vigencias), asi que la app no vuelve a filtrar nada de eso: lo que se
ve aqui es exactamente el universo del reporte.

Uso
---
    python -m streamlit run src/app_vigencias.py

Para publicarla gratis y que se abra con un enlace, ver PUBLICAR al final.
"""

import io
import os
from pathlib import Path

import altair as alt
import pandas as pd
import streamlit as st

RAIZ = Path(__file__).resolve().parent.parent
RUTA_DATOS = RAIZ / "output" / "COMPARACION_VIGENCIA_PUBLICO.parquet"
RUTA_DETALLE = RAIZ / "output" / "COMPARACION_VIGENCIA_DETALLE.parquet"

# El libro de revision que escribe comparacion_vigencia.py en cada corrida:
# results/COMPARACION_VIGENCIA/COMPARACION_VIGENCIA_<fecha>.xlsx. Es el mismo
# archivo que se descarga a mano hoy, con las ocho hojas del reporte, y desde
# la app se entrega TAL CUAL, sin volver a armarlo: lo que se revisa en Excel y
# lo que se baja de la app tienen que ser byte por byte el mismo documento.
CARPETA_REPORTE = RAIZ / "results" / "COMPARACION_VIGENCIA"
# Copia de nombre fijo que deja comparacion_vigencia.py al terminar. Es la UNICA
# que se versiona -results/ entero queda fuera, ver .gitignore- y por eso es la
# que tiene el deploy. En local suelen estar las dos y gana la mas nueva.
RUTA_REPORTE_FIJA = RAIZ / "output" / "COMPARACION_VIGENCIA_REPORTE.xlsx"


def reporte_mas_reciente():
    """
    El libro del reporte mas nuevo entre los dos sitios donde puede estar, o
    None si no hay ninguno: en local los COMPARACION_VIGENCIA_<fecha>.xlsx de
    results/, en el deploy la copia fija de output/.
    """
    if CARPETA_REPORTE.is_dir():
        # Los de results/ mandan: traen la fecha en el nombre, que es como se
        # identifica una corrida al pasarse el archivo entre personas.
        con_fecha = sorted(CARPETA_REPORTE.glob("COMPARACION_VIGENCIA_*.xlsx"),
                           key=lambda f: f.stat().st_mtime, reverse=True)
        if con_fecha:
            return con_fecha[0]
    return RUTA_REPORTE_FIJA if RUTA_REPORTE_FIJA.exists() else None

# Las dos vigencias salen del modulo que genero el parquet, para que la app no
# quede con los anos escritos a mano cuando alla se corran. Si el import falla
# -en el servidor puede no estar matplotlib, que arrastra comparacion_ofertas-
# se usan los mismos valores por defecto que trae aquel CONFIG.
try:
    from comparacion_vigencia import CONFIG as CONFIG_VIGENCIA
    V_BASE = CONFIG_VIGENCIA["vigencia_base"]
    V_LIQ = CONFIG_VIGENCIA["vigencia_liq"]
except Exception:                                            # pragma: no cover
    V_BASE, V_LIQ = 2026, 2027

# Los mismos seis cortes del reporte y del ejercicio 2025.
PERCENTILES = [10, 25, 50, 75, 90, 100]

# --- Formato de numeros -----------------------------------------------------
# Todo se muestra a la colombiana: miles con PUNTO y decimales con COMA. No se
# usa el formato "localized" de Streamlit ni el de los graficos porque esos
# siguen el idioma del NAVEGADOR de quien abre la app: el mismo numero le
# saldria 476.900 a uno y 476,900 a otro, que aqui son cosas distintas.
#
# En las tablas el formato se aplica con un Styler y no con column_config, para
# que la columna siga siendo numerica y se pueda seguir ordenando al hacer clic
# en el encabezado; column_config manda sobre el Styler, asi que donde se usa
# este no se le pone 'format' a aquel.
LOCALE_VEGA = {"number": {"decimal": ",", "thousands": ".", "grouping": [3],
                          "currency": ["$ ", ""]}}


def pesos(v, decimales: int = 0) -> str:
    """1234567.8 -> '$ 1.234.568'"""
    return "" if v is None or pd.isna(v) else f"$ {_miles(v, decimales)}"


def pct(v, decimales: int = 2, signo: bool = False) -> str:
    """6.9123 -> '6,91 %'  (con signo=True, '+6,91 %')"""
    if v is None or pd.isna(v):
        return ""
    texto = _miles(v, decimales)
    return f"+{texto} %" if signo and v > 0 else f"{texto} %"


def pesos_signo(v, decimales: int = 0) -> str:
    """
    Una diferencia SIEMPRE con su signo: '+$ 1.234.568' o '-$ 1.234.568'.

    El signo va delante del de pesos y no detras -'$ -1.234.568', que es lo que
    sale de formatear a secas-. El '+' de las positivas es a proposito: en una
    columna de diferencias lo que se busca de un vistazo es hacia donde se
    mueve cada fila, y sin el hay que fijarse en si la de al lado trae menos.
    """
    if v is None or pd.isna(v):
        return ""
    return ("+" if v >= 0 else "-") + pesos(abs(v), decimales)


def entero(v) -> str:
    """206875 -> '206.875'"""
    return "" if v is None or pd.isna(v) else _miles(v, 0)


def _miles(v: float, decimales: int) -> str:
    """
    El truco de siempre: Python solo sabe agrupar con coma, asi que se formatea
    a la inglesa y despues se intercambian los dos separadores. El paso por el
    espacio duro evita que el segundo replace deshaga el primero.
    """
    return (f"{v:,.{decimales}f}"
            .replace(",", " ").replace(".", ",").replace(" ", "."))

# La misma paleta de los PNG del reporte (VIZ en comparacion_ofertas.py):
# azul = liquidacion, naranja = el valor contra el que se compara.
AZUL, NARANJA = "#2a78d6", "#eb6834"

# Que se lee: tres MEDIDAS -el VM2 de la construccion, el valor total
# construido (ese VM2 por el area) y el avaluo del predio- en dos BASES
# (catastral, que es sobre lo que se cobra, o comercial, que es lo que se
# estima que vale). Son seis combinaciones y el parquet trae las seis, asi que
# cambiar de una a otra no vuelve a calcular nada pesado.
#
# El VM2 y el total construido se mueven igual en porcentaje -el area es la
# misma a los dos lados- pero no en pesos: el total dice cuanto vale el
# cambio, el VM2 solo a que precio quedo el metro.
#
# El comercial de la vigencia sale de dividir el catastral por el factor de la
# comuna: 0.7 en las actualizadas en 2024-2025 y 0.6 en el resto. Por eso la
# variacion comercial NO es igual a la catastral: la liquidacion se baja con un
# 0.7 parejo y la vigencia no.
MEDIDAS = {
    "Valor por m²": ("VM2", "Valor por m²"),
    "Valor total construido": ("VALORCONS", "Valor total construido"),
    "Avalúo": ("AVALUO", "Avalúo"),
}
BASES = {"Catastral": "CATASTRAL", "Comercial": "COMERCIAL"}

SERIES = {
    ("VALORCONS", "CATASTRAL"): {
        "vig": "VALORCONS_CAT_VIGENCIA", "liq": "VALORCONS_CAT_LIQ",
        "var": "VARIACION_VALORCONS_CAT_PCT", "prefijo": "VALORCONS_CAT",
        "dif": "DIF_VALORCONS_CAT"},
    ("VALORCONS", "COMERCIAL"): {
        "vig": "VALORCONS_COM_VIGENCIA", "liq": "VALORCONS_COM_LIQ",
        "var": "VARIACION_VALORCONS_COM_PCT", "prefijo": "VALORCONS_COM",
        "dif": "DIF_VALORCONS_COM"},
    ("VM2", "CATASTRAL"): {
        "vig": "VM2_CAT_VIGENCIA", "liq": "VM2_CAT_LIQ",
        "var": "VARIACION_CAT_PCT", "prefijo": "VM2_CAT",
        "dif": "DIF_CAT_ABS"},
    ("VM2", "COMERCIAL"): {
        "vig": "VM2_COM_VIGENCIA", "liq": "VM2_COM_LIQ",
        "var": "VARIACION_COM_PCT", "prefijo": "VM2_COM",
        "dif": "DIF_COM_ABS"},
    ("AVALUO", "CATASTRAL"): {
        "vig": "AVALUO_CAT_VIGENCIA", "liq": "AVALUO_CAT_LIQ",
        "var": "VARIACION_AVALUO_CAT_PCT", "prefijo": "AVALÚO_CAT",
        "dif": "DIF_AVALUO_CAT"},
    ("AVALUO", "COMERCIAL"): {
        "vig": "AVALUO_COM_VIGENCIA", "liq": "AVALUO_COM_LIQ",
        "var": "VARIACION_AVALUO_COM_PCT", "prefijo": "AVALÚO_COM",
        "dif": "DIF_AVALUO_COM"},
}

# --- Reglas de asignacion de tabla (hoja informativa) -----------------------
# Esto es la ESPECIFICACION, no un recuento de la corrida: describe como se
# asigna la tabla en general, sirva o no para los predios que hoy hay en el
# parquet. Es el recordatorio del metodo.
#
# Los dos grupos de comunas tienen que coincidir con COMUNAS_7 y COMUNAS_10 de
# tabla_construccion.py, que es donde se definen de verdad. Si alla se mueve
# una comuna de grupo -como paso al incluir 05, 06, 13, 16 y 18-, aqui hay que
# moverla tambien: son la misma regla escrita en dos sitios porque la app tiene
# que poder correr sin importar el pipeline.
GRUPOS_COMUNAS = {
    "7C": ["02", "03", "04", "08", "17", "19", "22"],
    "10C": ["01", "05", "06", "07", "09", "10", "11", "12", "13", "14", "15",
            "16", "18", "20", "21"],
}

# Los tres grupos con que se reparten las 22 comunas, que es como el filtro las
# ofrece. Los dos primeros son los de las tablas de valor (COMUNAS_7 y
# COMUNAS_10 de tabla_construccion.py); el tercero son las cinco que entraron
# despues -INCLUIR_COMUNAS_FALTANTES = True- y que hoy se liquidan leyendo esas
# mismas columnas *_10C_*, no una tabla propia.
GRUPOS_FILTRO = {
    "10 comunas": ["01", "07", "09", "10", "11", "12", "14", "15", "20", "21"],
    "7 comunas": ["02", "03", "04", "08", "17", "19", "22"],
    "5 comunas (extra)": ["05", "06", "13", "16", "18"],
}

# comuna -> grupo, para armar la columna con la que se filtra y se abre.
COMUNA_A_GRUPO = {c: g for g, cs in GRUPOS_FILTRO.items() for c in cs}

USOS_T1 = ("Casas (001), Barracas (004), Vivienda_Hasta_3_Pisos (012), "
           "Vivienda_Hasta_3_Pisos_En_PH (013), Jardin_Infantil_en_Casa (063)")
USOS_T2 = "Apartamentos_4_y_mas_pisos (003)"

# (uso, grupo, condicion juridica, patron de la columna). El patron lleva
# {t} donde va la tipologia de la ZHF.
REGLAS_TABLA = [
    (USOS_T1, "10C", "9",
     "T1_RESIDENCIAL_10C_COND_9_{t}"),
    (USOS_T1, "10C", "Diferente de 9", "T1_RESIDENCIAL_10C_COND_0_{t}"),
    (USOS_T1, "7C", "Todas", "T1_RESIDENCIAL_7C_{t}"),
    (USOS_T2, "10C", "Diferente de 8 y 9", "T2_EDIFICIOS_10C_{t}"),
    (USOS_T2, "7C", "Diferente de 8 y 9", "T2_EDIFICIOS_7C_{t}"),
]

TIPOLOGIAS_ZHF = ["011", "012", "013", "014", "015", "016"]


def reglas_asignacion() -> pd.DataFrame:
    """La especificacion desplegada: una fila por regla y tipologia."""
    filas = [{"USO DE CONSTRUCCIÓN": uso,
              "COMUNAS": ", ".join(GRUPOS_COMUNAS[grupo]),
              "CONDICIÓN JURÍDICA": condicion,
              "TIPOLOGÍA ZHF": t,
              "REFERENCIA TABLA": patron.format(t=t)}
             for uso, grupo, condicion, patron in REGLAS_TABLA
             for t in TIPOLOGIAS_ZHF]
    return pd.DataFrame(filas)


# Por que columna se parten el resumen, los percentiles y los graficos: se
# arma una fila -y un grafico- por cada valor distinto de esa columna.
#
# "Actualizacion 2024-2025" salio de aqui: partia las 22 comunas en las mismas
# dos mitades que "Grupo de comunas" pero con otro nombre, y en pantalla las
# dos opciones juntas no se distinguian.
APERTURAS = {
    "Tabla de valor": "TABLA_ORIGEN",
    "Comuna": "COMUNA",
    "Actividad económica de la ZHF": "ACTIVIDAD_ECONOMICA",
    "Grupo de comunas": "GRUPO_COMUNAS",
    "Tabla y actividad juntas (como en el reporte)": "CLAVE",
}

# Todo lo que trae el parquet publico. Si alguna vez hay que sumar una columna,
# revisar primero que no permita senalar a un predio en concreto.
COLUMNAS = ["COMUNA", "ACTUALIZACION", "TABLA_ORIGEN", "TABLA_VALOR",
            "USO_LADM", "ACTIVIDAD_ECONOMICA", "CLAVE",
            "VALORCONS_CAT_VIGENCIA", "VALORCONS_CAT_LIQ",
            "VARIACION_VALORCONS_CAT_PCT",
            "VALORCONS_COM_VIGENCIA", "VALORCONS_COM_LIQ",
            "VARIACION_VALORCONS_COM_PCT",
            "VM2_CAT_VIGENCIA", "VM2_CAT_LIQ", "VARIACION_CAT_PCT",
            "AVALUO_CAT_VIGENCIA", "AVALUO_CAT_LIQ",
            "VARIACION_AVALUO_CAT_PCT",
            "VM2_COM_VIGENCIA", "VM2_COM_LIQ", "VARIACION_COM_PCT",
            "AVALUO_COM_VIGENCIA", "AVALUO_COM_LIQ", "VARIACION_AVALUO_COM_PCT"]


st.set_page_config(page_title=f"Comparación de vigencias {V_BASE} → {V_LIQ}",
                   page_icon="🏙️", layout="wide")

st.markdown(
    """
    <style>
        /* Compacto a proposito: el encabezado y las tarjetas se comian media
           pantalla al abrir, y lo que se viene a ver son las tablas. Todo lo
           de arriba tiene que caber en una franja y dejar el contenido a la
           vista sin desplazarse. */
        .block-container {padding-top: .8rem; padding-bottom: 1.2rem;}
        div[data-testid="stMetric"] {
            background:#ffffff; border:1px solid #e6e9ef; border-radius:9px;
            padding:7px 11px; box-shadow:0 1px 2px rgba(16,24,40,.05);
        }
        div[data-testid="stMetricLabel"] p {
            color:#667085; font-weight:600; font-size:12px;
        }
        div[data-testid="stMetricValue"] {font-size:20px; line-height:1.2;}
        div[data-testid="stMetricDelta"] {font-size:11px;}
        .app-hero {
            background:linear-gradient(120deg,#1e3a8a 0%,#2563eb 55%,#0ea5e9 100%);
            color:#fff; padding:11px 16px; border-radius:10px; margin-bottom:10px;
        }
        .app-hero h1 {margin:0; font-size:18px; font-weight:700;}
        .app-hero p  {margin:2px 0 0; opacity:.88; font-size:12.5px;}
    </style>
    """,
    unsafe_allow_html=True,
)


# =====================================================================
# DATOS
# =====================================================================
@st.cache_data(show_spinner="Leyendo el detalle de la comparación…")
def cargar(ruta: str, marca_tiempo: float) -> pd.DataFrame:
    """
    El detalle, con solo las columnas que usa la app.

    marca_tiempo es la fecha del archivo: no se usa adentro, pero al entrar en
    la firma hace que el cache se invalide solo cuando se vuelve a correr
    comparacion_vigencia.py, sin tener que reiniciar la app.
    """
    try:
        import pyarrow.parquet as pq
        hay = set(pq.ParquetFile(ruta).schema_arrow.names)
        d = pd.read_parquet(ruta, columns=[c for c in COLUMNAS if c in hay])
    except ImportError:                                      # pragma: no cover
        d = pd.read_parquet(ruta)
        d = d[[c for c in COLUMNAS if c in d.columns]]
    d["COMUNA"] = d["COMUNA"].astype(str).str.strip().str.zfill(2)
    for c in ("TABLA_ORIGEN", "TABLA_VALOR", "USO_LADM",
              "ACTIVIDAD_ECONOMICA", "CLAVE"):
        if c in d.columns:
            d[c] = d[c].astype(str)
    # Grupo de comunas: es con lo que se filtra y se abre, no viene en el parquet.
    d["GRUPO_COMUNAS"] = (d["COMUNA"].map(COMUNA_A_GRUPO)
                          .fillna("sin grupo"))
    return d


@st.cache_data(show_spinner="Leyendo el detalle predio a predio…")
def cargar_detalle(ruta: str, marca_tiempo: float) -> pd.DataFrame:
    """
    El detalle fila a fila, CON identificadores. Solo lo lee la hoja Detalle.

    Son 288 mil filas por 50 columnas: ~345 MB en memoria. Se cachea igual que
    el publico y con la misma marca de tiempo, asi que se lee una sola vez por
    corrida del pipeline aunque se cambie de filtro cien veces.
    """
    d = pd.read_parquet(ruta)
    d["COMUNA"] = d["COMUNA"].astype(str).str.strip().str.zfill(2)
    for c in ("ID_PREDIO", "NUMERO_PREDIAL_NACIONAL", "CONSTRUCCION_ID",
              "TABLA_ORIGEN", "TABLA_VALOR", "USO_LADM", "ZHF",
              "ACTIVIDAD_ECONOMICA", "CLAVE", "SENTIDO", "RANGO_VARIACION"):
        if c in d.columns:
            d[c] = d[c].astype(str)
    d["GRUPO_COMUNAS"] = d["COMUNA"].map(COMUNA_A_GRUPO).fillna("sin grupo")
    return d


# Las columnas del detalle que se muestran en la vista corta. Las que dependen
# de la medida elegida -las cuatro de valor- se agregan despues, ya sabiendo
# cual pidio el usuario en la barra lateral.
COLUMNAS_DETALLE_FIJAS = [
    "ID_PREDIO", "NUMERO_PREDIAL_NACIONAL", "CONSTRUCCION_ID", "COMUNA",
    "GRUPO_COMUNAS", "ESTRPRED", "USO_LADM", "CONDICION", "TABLA_ORIGEN",
    "TABLA_VALOR", "ZHF", "ACTIVIDAD_ECONOMICA", "PUNTCONS", "AREA_CONST",
]

# Que es cada columna del detalle. Se trae de comparacion_vigencia.py, que es
# donde esta escrito de verdad; si el import falla -en el servidor puede no
# estar matplotlib- la hoja Diccionario simplemente no sale.
try:
    from comparacion_vigencia import DICCIONARIO_DETALLE
except Exception:                                        # pragma: no cover
    DICCIONARIO_DETALLE = []


if not os.path.exists(RUTA_DATOS):
    st.error(
        f"No se encontró **{RUTA_DATOS.name}**.\n\n"
        "Corra primero `python src/comparacion_vigencia.py`, que es quien lo "
        "escribe en `output/`."
    )
    st.stop()

df = cargar(str(RUTA_DATOS), os.path.getmtime(RUTA_DATOS))

st.markdown(
    f"""
    <div class="app-hero">
        <h1>🏙️ Comparación de vigencias · {V_BASE} → {V_LIQ}</h1>
        <p>La liquidación ({V_LIQ}) contra lo que cobra hoy la base
        ({V_BASE}) · <b>{entero(len(df))} predios</b> de una sola construcción
        con valor de tabla en las dos vigencias.</p>
    </div>
    """,
    unsafe_allow_html=True,
)


# =====================================================================
# FILTROS (uno solo para las tres hojas)
# =====================================================================
with st.sidebar:
    st.header("⚙️ Filtros")
    st.caption("Se aplican a las tres hojas. Vacío = todo.")

    etiqueta_medida = st.radio("Medida", list(MEDIDAS), index=0)
    etiqueta_base = st.radio(
        "Base de valor", list(BASES), index=0, horizontal=True,
        help="Catastral es lo que se cobra. Comercial es lo que se estima que "
             "vale: el catastral de la vigencia dividido por 0,7 en las comunas "
             "actualizadas en 2024-2025 y por 0,6 en las demás.")
    clave_medida, titulo_medida = MEDIDAS[etiqueta_medida]
    medida = SERIES[(clave_medida, BASES[etiqueta_base])]
    unidad_eje = f"{titulo_medida} {etiqueta_base.lower()} (millones de pesos)"

    familias = sorted({t.split("_")[0] + "_" + t.split("_")[1]
                       for t in df["TABLA_ORIGEN"].unique() if "_" in t})
    sel_familia = st.multiselect("Categoría de tabla", familias)

    # Las tablas que se ofrecen dependen de la familia elegida, para no dar a
    # escoger entre 40 codigos cuando ya se acoto a residencial.
    de_la_familia = (df["TABLA_ORIGEN"].str.startswith(tuple(sel_familia))
                     if sel_familia else slice(None))
    tablas = sorted(df.loc[de_la_familia, "TABLA_ORIGEN"].unique())
    sel_tabla = st.multiselect("Tabla de valor", tablas)

    sel_comuna = st.multiselect("Comuna", sorted(df["COMUNA"].unique()))
    sel_actividad = st.multiselect("Actividad económica de la ZHF",
                                   sorted(df["ACTIVIDAD_ECONOMICA"].unique()))
    sel_grupo = st.multiselect(
        "Grupo de comunas", list(GRUPOS_FILTRO),
        help="Qué comunas trae cada grupo está en la hoja Reglas.")

    st.divider()
    etiqueta_apertura = st.selectbox(
        "Separar los resultados por", list(APERTURAS), index=0,
        help="Con lo que se elija aquí se parten las tablas y los gráficos: "
             "una fila del resumen y un gráfico por cada valor de esa "
             "columna. Con «Comuna», por ejemplo, sale una fila y un gráfico "
             "por cada comuna.")
    col_apertura = APERTURAS[etiqueta_apertura]
    min_predios = st.number_input(
        "Mínimo de predios por grupo", 1, 5000, 5,
        help="Los grupos con menos predios de los que se pidan aquí no se "
             "muestran: con tan pocos casos una mediana no dice nada.")

def filtrar(d: pd.DataFrame) -> pd.DataFrame:
    """
    Los filtros de la barra lateral, aplicados a lo que se le pase.

    Esta factorizado porque lo usan los dos parquet: el publico que mueve las
    hojas de agregados y el detalle que mueve la hoja de predios. Si el filtro
    viviera escrito dos veces, tarde o temprano una hoja quedaria mostrando
    algo distinto de la otra con la misma seleccion en pantalla.
    """
    if sel_familia:
        d = d[d["TABLA_ORIGEN"].str.startswith(tuple(sel_familia))]
    if sel_tabla:
        d = d[d["TABLA_ORIGEN"].isin(sel_tabla)]
    if sel_comuna:
        d = d[d["COMUNA"].isin(sel_comuna)]
    if sel_actividad:
        d = d[d["ACTIVIDAD_ECONOMICA"].isin(sel_actividad)]
    if sel_grupo:
        d = d[d["GRUPO_COMUNAS"].isin(sel_grupo)]
    # La medida de avaluo puede venir vacia si se corrio la comparacion sin las
    # columnas de terreno; mejor decirlo que mostrar una hoja en blanco.
    return d[d[medida["vig"]].notna() & d[medida["liq"]].notna()]


dff = filtrar(df)

if dff.empty:
    st.warning("Ningún predio cumple los filtros elegidos. Quite alguno en la "
               "barra lateral.")
    st.stop()

c_base, c_liq, c_var = (f"{medida['prefijo']}_VIG_{V_BASE}",
                        f"{medida['prefijo']}_VIG_{V_LIQ}",
                        f"VARIACIÓN_{medida['prefijo']}_{V_LIQ}_vs_{V_BASE}")
# La diferencia absoluta -en pesos- entre las dos vigencias. La variacion de esa
# diferencia contra la vigencia base es justamente c_var, asi que las dos
# columnas van siempre juntas: cuanto cambia y en que proporcion.
c_dif = f"DIFERENCIA_{medida['prefijo']}_{V_LIQ}_vs_{V_BASE}"

k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("Predios", entero(len(dff)),
          f"{pct(len(dff) / len(df) * 100, 1)} del total")
k2.metric(f"Mediana vigencia {V_BASE}", pesos(dff[medida["vig"]].median()))
k3.metric(f"Mediana vigencia {V_LIQ}", pesos(dff[medida["liq"]].median()))
k4.metric("Variación mediana", pct(dff[medida["var"]].median(), 2, signo=True))
k5.metric("Bajan", pct((dff[medida["var"]] < 0).mean() * 100, 1))


# =====================================================================
# CALCULOS (los mismos que arma comparacion_vigencia.py)
# =====================================================================
def percentiles(s: pd.DataFrame) -> pd.DataFrame:
    """
    Los seis cortes de un grupo, con las dos series y su variacion.

    Cada serie se ordena POR SEPARADO, igual que en el Excel: el p90 de una
    vigencia y el p90 de la otra no son el mismo predio, asi que esto compara
    distribuciones, no casos. La variacion pareada -predio contra si mismo- es
    la que sale en los KPI de arriba y en la hoja de graficos.
    """
    v = pd.to_numeric(s[medida["vig"]], errors="coerce").dropna()
    l = pd.to_numeric(s[medida["liq"]], errors="coerce").dropna()
    if v.empty or l.empty:
        return pd.DataFrame()
    filas = []
    for p in PERCENTILES:
        pv, pl = float(v.quantile(p / 100)), float(l.quantile(p / 100))
        filas.append({"PERCENTIL": f"{p}%",
                      "NUM_PREDIOS": int(round(p / 100 * len(s))),
                      c_base: pv, c_liq: pl,
                      c_dif: pl - pv,
                      c_var: (pl / pv - 1) if pv else None})
    return pd.DataFrame(filas)


def por_grupo(d: pd.DataFrame, col: str) -> pd.DataFrame:
    """Los percentiles de cada grupo, uno debajo de otro y en una sola tabla."""
    salida = []
    for clave, s in d.groupby(col, sort=True):
        if len(s) < min_predios:
            continue
        t = percentiles(s)
        if t.empty:
            continue
        t.insert(0, etiqueta_apertura, str(clave))
        salida.append(t)
    return pd.concat(salida, ignore_index=True) if salida else pd.DataFrame()


def con_formato(t: pd.DataFrame):
    """
    La tabla lista para mostrar: pesos en las series, porcentaje en las
    variaciones y miles en los conteos, todo a la colombiana.

    Devuelve un Styler, no un DataFrame: asi la columna sigue siendo numerica
    -se puede ordenar haciendo clic en el encabezado- y lo unico que cambia es
    como se dibuja. La variacion de los percentiles viene en FRACCION (0.0691)
    y la de los resumenes en PUNTOS (6.91), por eso el x100.
    """
    reglas = {}
    for col in t.columns:
        if col == c_dif:                         # pesos, con signo
            reglas[col] = pesos_signo
        elif col in (c_base, c_liq):
            reglas[col] = pesos
        elif col == c_var:                       # fraccion, con signo
            reglas[col] = lambda v: pct(v * 100 if pd.notna(v) else v, 2,
                                        signo=True)
        elif col == "VAR_MEDIANA_%":             # puntos, con signo
            reglas[col] = lambda v: pct(v, 2, signo=True)
        elif col.endswith("_%"):                 # BAJAN_% y SUBEN_% son
            reglas[col] = lambda v: pct(v, 2)    # proporciones, no variaciones
        elif col in ("NUM_PREDIOS", "PREDIOS"):
            reglas[col] = entero
    return t.style.format(reglas)


def resumen(d: pd.DataFrame, col: str) -> pd.DataFrame:
    """
    Una fila por grupo: cuantos predios, las dos medianas y como se mueve.

    Es la hoja General del reporte, pero sobre lo filtrado y por la columna que
    se elija. Todo son agregados de por lo menos min_predios predios: ninguna
    fila describe a un predio en particular.
    """
    g = d.groupby(col, sort=True)
    t = pd.DataFrame({
        "PREDIOS": g.size(),
        c_base: g[medida["vig"]].median(),
        c_liq: g[medida["liq"]].median(),
        "VAR_MEDIANA_%": g[medida["var"]].median(),
        "BAJAN_%": g[medida["var"]].apply(lambda s: (s < 0).mean() * 100),
        "SUBEN_%": g[medida["var"]].apply(lambda s: (s > 0).mean() * 100),
    })
    t = t[t["PREDIOS"] >= min_predios]
    if t.empty:
        return pd.DataFrame()

    grupos_ok = list(t.index)                    # antes de perder el indice
    t.insert(3, c_dif, t[c_liq] - t[c_base])     # queda detras de las dos medianas
    t = t.reset_index().rename(columns={col: etiqueta_apertura})

    # Fila de totales sobre LOS MISMOS grupos que quedaron en la tabla, para que
    # PREDIOS sea exactamente la suma de la columna. No es la suma de las demas
    # columnas: una mediana no se suma, se vuelve a calcular sobre el conjunto.
    sub = d[d[col].isin(grupos_ok)]
    fila = {etiqueta_apertura: "TOTAL",
            "PREDIOS": len(sub),
            c_base: sub[medida["vig"]].median(),
            c_liq: sub[medida["liq"]].median(),
            "VAR_MEDIANA_%": sub[medida["var"]].median(),
            "BAJAN_%": (sub[medida["var"]] < 0).mean() * 100,
            "SUBEN_%": (sub[medida["var"]] > 0).mean() * 100}
    fila[c_dif] = fila[c_liq] - fila[c_base]
    return pd.concat([t, pd.DataFrame([fila])[t.columns]], ignore_index=True)


hoja_tablas, hoja_graf, hoja_detalle, hoja_reglas = st.tabs(
    ["📊 Tablas", "📈 Gráficos", "🔎 Detalle",
     "📋 Reglas"])

# Todo se arma una sola vez aca arriba: las tres hojas y el Excel leen de estos
# mismos objetos, asi no hay dos sitios calculando lo mismo y desviandose.
res = resumen(dff, col_apertura)
total = percentiles(dff)
abierto = por_grupo(dff, col_apertura)
reglas = reglas_asignacion()

# El reparto de la variacion. Aqui si es predio contra si mismo: cada uno cae
# en el rango de su propia variacion, no se comparan distribuciones.
RANGOS = ["Baja más de 50%", "Baja 25-50%", "Baja 10-25%", "Estable (±10%)",
          "Sube 10-25%", "Sube 25-50%", "Sube más de 50%"]
reparto = (dff[medida["var"]]
           .pipe(pd.cut,
                 bins=[-float("inf"), -50, -25, -10, 10, 25, 50, float("inf")],
                 labels=RANGOS)
           .value_counts(sort=False)
           .rename_axis("RANGO").reset_index(name="PREDIOS"))
reparto["%"] = reparto["PREDIOS"] / reparto["PREDIOS"].sum() * 100


def _motor_excel():
    """
    El primer motor de Excel que este instalado, o None si no hay ninguno.

    En Streamlit Community Cloud las dependencias salen de requirements.txt,
    pero agregar una linea alli NO siempre dispara la reinstalacion: si el
    entorno quedo del deploy anterior, no hay xlsxwriter ni openpyxl. Antes eso
    reventaba la pagina COMPLETA -las tres hojas en blanco con un
    ModuleNotFoundError-, porque el libro se arma en cada render aunque se este
    mirando Graficos o Reglas. Ahora se detecta antes y se baja a CSV.
    """
    for modulo in ("xlsxwriter", "openpyxl"):
        try:
            __import__(modulo)
            return modulo
        except ImportError:
            continue
    return None


MOTOR_EXCEL = _motor_excel()


def excel_predios(d: pd.DataFrame) -> bytes:
    """
    El detalle fila a fila a Excel, para revisar casos a mano.

    Es el mismo libro que escribe comparacion_vigencia.py en results/, pero
    armado sobre lo que este filtrado en pantalla y en memoria, sin pasar por
    disco. Dos hojas:

        Detalle       una fila por construccion, con autofiltro y las dos
                      primeras columnas congeladas para que el identificador no
                      se pierda al desplazarse a la derecha
        Diccionario   que es cada columna y de donde sale, para que quien abra
                      el archivo no tenga que venir a leer el codigo

    Sin constant_memory a proposito: ese modo obliga a escribir en orden de
    fila y pandas emite columna por columna, asi que la hoja saldria en blanco.
    Por eso la hoja Detalle va con tope de filas, no con el universo completo.
    """
    d = d.copy()
    for c in d.columns:
        if isinstance(d[c].dtype, pd.CategoricalDtype):
            d[c] = d[c].astype(str)

    dic = pd.DataFrame(
        [(col, desc, nota) for col, desc, nota in DICCIONARIO_DETALLE
         if col in d.columns],
        columns=["COLUMNA", "QUÉ ES", "DE DÓNDE SALE"])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine=MOTOR_EXCEL) as xw:
        libro = xw.book if MOTOR_EXCEL == "xlsxwriter" else None
        if libro is not None:
            f_tit = libro.add_format({"bold": True, "bg_color": "#1F4E78",
                                      "font_color": "white", "border": 1,
                                      "align": "center", "valign": "vcenter",
                                      "text_wrap": True})
            f_pesos = libro.add_format({"num_format": "$ #,##0"})
            f_pct = libro.add_format({"num_format": "0.00"})
            f_ent = libro.add_format({"num_format": "#,##0"})
            f_texto = libro.add_format({"text_wrap": True, "valign": "top"})

        for nombre, t in (("Detalle", d), ("Diccionario", dic)):
            if t.empty:
                continue
            t.to_excel(xw, sheet_name=nombre, index=False)
            if libro is None:
                continue
            h = xw.sheets[nombre]
            h.freeze_panes(1, 2 if nombre == "Detalle" else 0)
            h.autofilter(0, 0, len(t), len(t.columns) - 1)
            for i, col in enumerate(t.columns):
                h.write(0, i, str(col), f_tit)
                nom = str(col)
                if nombre == "Diccionario":
                    fmt, ancho = f_texto, (26 if i == 0 else 60)
                elif nom.startswith(("VM2", "VALORCONS", "AVALUO", "AVALPRED",
                                     "VTER", "VANEXO", "DIF_")):
                    fmt, ancho = f_pesos, 20
                elif nom.startswith("VARIACION") or nom.endswith("_PCT"):
                    fmt, ancho = f_pct, 16
                elif nom in ("PUNTCONS", "ACONCONS", "AREA_CONST", "ESTRPRED",
                             "CONDICION", "N_CONST_PREDIO"):
                    fmt, ancho = f_ent, 14
                else:
                    largo = int(t[col].astype(str).str.len().head(500).max() or 12)
                    fmt, ancho = None, min(max(14, largo + 3, len(nom) + 3), 40)
                h.set_column(i, i, ancho, fmt)
    return buf.getvalue()


# ---------------------------------------------------------------------
# HOJA 1 - TABLAS
# ---------------------------------------------------------------------
with hoja_tablas:
    st.subheader(f"Tablas · {etiqueta_medida} · base {etiqueta_base.lower()}")

    # Sin boton de descarga aqui: la unica descarga de la app esta en la hoja
    # Detalle, y es el libro del reporte. Tener dos era pedirle a quien la abre
    # que adivinara cual de los dos archivos es "el bueno".

    st.markdown(f"**Resumen por {etiqueta_apertura.lower()}**")
    st.caption(f"Comparación de los valores **medianos** de cada grupo entre "
               f"las vigencias {V_BASE} y {V_LIQ}.")
    if res.empty:
        st.info(f"Ningún grupo llega a {min_predios} predios con estos filtros.")
    else:
        # El matiz de por que DIFERENCIA y VAR_MEDIANA_% no siempre concuerdan
        # -una compara distribuciones y la otra es predio contra si mismo- pasa
        # a la ayuda de cada encabezado: quien lo necesite lo tiene al pasar el
        # mouse, y el resto no lee cinco renglones antes de ver la tabla.
        #
        # OJO: aqui column_config va SIN 'format'. El formato lo pone el Styler
        # de con_formato() y column_config le ganaria, dejando los numeros en
        # crudo.
        st.dataframe(
            con_formato(res), width="stretch", hide_index=True,
            column_config={
                "PREDIOS": st.column_config.Column(
                    help="Construcciones del grupo que entran a la comparación."),
                c_base: st.column_config.Column(
                    help=f"Valor mediano del grupo en la vigencia {V_BASE}: la "
                         f"mitad de los predios está por debajo."),
                c_liq: st.column_config.Column(
                    help=f"Valor mediano del grupo con la liquidación "
                         f"({V_LIQ}), calculado igual."),
                c_dif: st.column_config.Column(
                    help="La mediana de {} menos la de {}. Compara las dos "
                         "distribuciones, no predio contra predio: el predio "
                         "que queda en el medio en una vigencia no tiene por "
                         "qué ser el mismo de la otra.".format(V_LIQ, V_BASE)),
                "VAR_MEDIANA_%": st.column_config.Column(
                    help="La mediana de las variaciones, y esta SÍ es predio "
                         "contra sí mismo. Por eso no coincide con la "
                         "diferencia dividida entre el valor base: son dos "
                         "cuentas distintas, no un error."),
                "BAJAN_%": st.column_config.Column(
                    help="Qué proporción del grupo baja de una vigencia a otra."),
                "SUBEN_%": st.column_config.Column(
                    help="Qué proporción del grupo sube de una vigencia a otra."),
            })
        st.caption("La fila **TOTAL** consolida todos los grupos: *PREDIOS* es la suma, mientras que las medianas y porcentajes se recalculan sobre el conjunto total es sobre los mismos grupos que se ven ")

    st.divider()
    st.markdown("**Percentiles, diferencia y variación** · "
                f"{entero(len(dff))} predios")
    st.caption(f"En cada corte: cuánto vale en la vigencia {V_BASE}, cuánto "
               f"valdría en la {V_LIQ}, cuántos pesos de diferencia hay entre "
               f"los dos y qué proporción representa esa diferencia.")
    st.dataframe(con_formato(total), width="stretch", hide_index=True)



# ---------------------------------------------------------------------
# HOJA 3 - GRAFICOS
# ---------------------------------------------------------------------
with hoja_graf:
    st.subheader(f"Gráficos · {etiqueta_medida} · base {etiqueta_base.lower()}")

    def pie(t: pd.DataFrame) -> list:
        """Bajo el titulo va SOLO el conteo: es corto y nunca se corta."""
        if t.empty:
            return []
        return [f"{entero(int(t['NUM_PREDIOS'].iloc[-1]))} predios comparados"]

    def frase(t: pd.DataFrame) -> str:
        """
        La lectura del grafico en una frase, para el caption de DEBAJO.

        Va fuera del titulo a proposito: Vega mide el subtitulo contra el ancho
        del grafico y lo corta con puntos suspensivos -"son +$ 66.500 (..."-,
        que es justo donde estaba el dato. Un caption de Streamlit es texto
        normal, se acomoda en los renglones que necesite y no pierde nada.

        Y dice QUE es el porcentaje. "+29,1 %" a secas no se explica solo:
        es la diferencia medida contra el valor de la vigencia base.
        """
        fila = t[t["PERCENTIL"] == "50%"] if not t.empty else t
        if fila.empty or pd.isna(fila[c_var].iloc[0]):
            return ""
        pv, pl = float(fila[c_base].iloc[0]), float(fila[c_liq].iloc[0])
        var = float(fila[c_var].iloc[0]) * 100
        return (f"En el medio de la distribución (percentil 50) el valor "
                f"{'sube' if pl >= pv else 'baja'} de **{pesos(pv)}** en "
                f"{V_BASE} a **{pesos(pl)}** en {V_LIQ}. La diferencia es de "
                f"**{pesos_signo(pl - pv)}**, que sobre el valor de {V_BASE} "
                f"equivale a **{pct(var, 1, signo=True)}**.")

    def curvas(t: pd.DataFrame, titulo: str) -> alt.Chart:
        """
        Las dos curvas de percentiles, en millones y con los colores del reporte.

        En el aviso del punto van, ademas del valor de la serie, la DIFERENCIA
        en pesos entre las dos vigencias en ese mismo percentil y que proporcion
        representa. Como la diferencia es una sola por percentil, se mapea desde
        la tabla ancha y queda igual en los dos puntos de la columna: parado en
        cualquiera de los dos se lee lo mismo.
        """
        largo = t.melt(id_vars="PERCENTIL", value_vars=[c_base, c_liq],
                       var_name="Serie", value_name="Valor")
        largo["Millones"] = largo["Valor"] / 1e6
        largo["Etiqueta"] = largo["Valor"].map(pesos)
        ref = t.set_index("PERCENTIL")
        largo["Diferencia"] = largo["PERCENTIL"].map(
            ref[c_dif].map(pesos_signo))
        largo["Variacion"] = largo["PERCENTIL"].map(
            ref[c_var].map(lambda v: pct(v * 100, 1, signo=True)
                           if pd.notna(v) else ""))
        return (
            alt.Chart(largo,
                      title=alt.TitleParams(text=titulo, subtitle=pie(t),
                                            subtitleColor="#667085",
                                            anchor="start"))
            .mark_line(point=alt.OverlayMarkDef(size=70), strokeWidth=2.5)
            .encode(
                # labelAngle=0: los rotulos del eje x van en horizontal.
                x=alt.X("PERCENTIL:N", title="Percentil",
                        sort=[f"{p}%" for p in PERCENTILES],
                        axis=alt.Axis(labelAngle=0)),
                y=alt.Y("Millones:Q", title=unidad_eje),
                color=alt.Color("Serie:N", title=None,
                                scale=alt.Scale(domain=[c_base, c_liq],
                                                range=[NARANJA, AZUL]),
                                legend=alt.Legend(orient="top")),
                tooltip=[alt.Tooltip("PERCENTIL:N", title="Percentil"),
                         alt.Tooltip("Serie:N", title="Serie"),
                         alt.Tooltip("Etiqueta:N", title="Valor"),
                         alt.Tooltip("Diferencia:N",
                                     title=f"Diferencia {V_LIQ} vs {V_BASE}"),
                         alt.Tooltip("Variacion:N", title="Variación")],
            )
            .properties(height=340)
            .configure(locale=LOCALE_VEGA)      # ejes en formato colombiano
        )

    st.altair_chart(curvas(total, "Total de la selección"), width="stretch")
    st.caption(frase(total))

    st.divider()
    st.markdown(f"**Por {etiqueta_apertura.lower()}**")
    if abierto.empty:
        st.info(f"Ningún grupo llega a {min_predios} predios con estos filtros.")
    else:
        grupos = list(abierto[etiqueta_apertura].unique())
        # Con muchos grupos la grilla se vuelve ilegible; se muestran de a pocos
        # y el usuario elige cuales, que es justo lo que en el Excel no se puede.
        elegidos = st.multiselect(f"{etiqueta_apertura} a dibujar", grupos,
                                  default=grupos[:6])
        columnas_grilla = st.radio("Gráficos por fila", [1, 2, 3], index=1,
                                   horizontal=True)
        for i in range(0, len(elegidos), columnas_grilla):
            for col, nombre in zip(st.columns(columnas_grilla),
                                   elegidos[i:i + columnas_grilla]):
                t = abierto[abierto[etiqueta_apertura] == nombre]
                col.altair_chart(curvas(t, nombre), width="stretch")
                col.caption(frase(t))

    st.divider()
    st.markdown("**Cómo se reparte la variación**")
    st.caption("Aquí sí es predio contra sí mismo, no distribución contra "
               "distribución: cada predio cae en el rango de su propia "
               "variación.")
    st.altair_chart(
        alt.Chart(reparto).mark_bar(color=AZUL, cornerRadiusEnd=3).encode(
            x=alt.X("RANGO:N", title=None, sort=list(reparto["RANGO"]),
                    axis=alt.Axis(labelAngle=0)),   # rotulos en horizontal
            y=alt.Y("PREDIOS:Q", title="Predios"),
            tooltip=[alt.Tooltip("RANGO:N", title="Rango"),
                     alt.Tooltip("PREDIOS:Q", title="Predios", format=",.0f"),
                     alt.Tooltip("%:Q", title="% del total", format=",.1f")],
        ).properties(height=300).configure(locale=LOCALE_VEGA),
        width="stretch")



# ---------------------------------------------------------------------
# HOJA 3 - DETALLE DE LA LIQUIDACION
# ---------------------------------------------------------------------
# Dos cosas distintas, y conviene no confundirlas:
#
#   1. El LIBRO DEL REPORTE (results/COMPARACION_VIGENCIA_<fecha>.xlsx). Es el
#      mismo archivo que hoy se baja a mano, con las ocho hojas de siempre. Se
#      entrega tal cual, sin volver a armarlo: pesa 47 KB y no lleva ni
#      ID_PREDIO ni numero predial, asi que puede salir publicado sin problema.
#
#   2. El EXPLORADOR fila a fila, que vive de
#      COMPARACION_VIGENCIA_DETALLE.parquet. Ese SI lleva identificadores y el
#      .gitignore lo deja fuera del repositorio, asi que en local funciona y en
#      el deploy publico no esta y se dice, en vez de reventar.
with hoja_detalle:
    st.subheader("Detalle de la liquidación")

    @st.cache_data(show_spinner=False)
    def hoja_general(ruta: str, marca_tiempo: float):
        """
        Solo la hoja General del libro del reporte.

        Se abre el libro y se BORRAN las demas hojas, en vez de leer los datos
        y volver a escribirlos: asi la General llega con el mismo formato que
        tiene en el archivo original -encabezado, anchos, panel congelado- y no
        con el que se le ponga aqui. El resto de hojas se quedan afuera porque
        el proceso las produce para revisar a fondo, y lo que se pide desde la
        app es la General y nada mas.

        Devuelve (bytes, hojas_quitadas). Si openpyxl no esta o el libro no
        trae una hoja llamada General, entrega el archivo completo y lo dice.
        """
        try:
            from openpyxl import load_workbook
            libro_x = load_workbook(ruta)
            if "General" not in libro_x.sheetnames:
                raise KeyError("General")
            quitadas = [h for h in libro_x.sheetnames if h != "General"]
            for h in quitadas:
                del libro_x[h]
            buf = io.BytesIO()
            libro_x.save(buf)
            return buf.getvalue(), quitadas
        except Exception:                                # pragma: no cover
            with open(ruta, "rb") as f:
                return f.read(), None

    libro = reporte_mas_reciente()
    if libro is None:
        st.info(
            "**No se encontró el libro del reporte.** Se busca el "
            "`COMPARACION_VIGENCIA_<fecha>.xlsx` más reciente en "
            f"`{CARPETA_REPORTE.relative_to(RAIZ)}`. Lo escribe "
            "`python src/comparacion_vigencia.py` al terminar.")
    else:
        datos_general, quitadas = hoja_general(str(libro),
                                              libro.stat().st_mtime)
        st.markdown("**La hoja General de la liquidación**")
        st.download_button(
            "📗 Ver el detalle de la liquidación",
            data=datos_general,
            # La copia fija no lleva fecha en el nombre; se la pone la del
            # archivo, para que el que se baja del deploy se pueda identificar
            # igual que el que se saca a mano.
            file_name=(libro.name if libro != RUTA_REPORTE_FIJA else
                       f"COMPARACION_VIGENCIA_"
                       f"{pd.Timestamp(libro.stat().st_mtime, unit='s'):%Y%m%d}"
                       f".xlsx"),
            mime=("application/vnd.openxmlformats-officedocument"
                  ".spreadsheetml.sheet"),
            type="primary", key="dl_reporte",
            help="La hoja General del libro que produce la liquidación, con "
                 "su formato original.")
        st.caption(
            f"Sale de `{libro.name}` · "
            f"{_miles(len(datos_general) / 1024, 0)} KB · generado el "
            f"{pd.Timestamp(libro.stat().st_mtime, unit='s'):%Y-%m-%d %H:%M}"
            + (f" · se dejaron fuera las otras {len(quitadas)} hojas "
               f"({', '.join(quitadas)})" if quitadas else
               " · ATENCIÓN: no se pudo recortar, va el libro completo")
            + ". No responde a los filtros de la izquierda: es la General de "
              "la corrida, igual a la que produce el proceso.")

    st.divider()
    st.markdown("**Explorador predio a predio**")

    if not os.path.exists(RUTA_DETALLE):
        st.info(
            f"**{RUTA_DETALLE.name} no está en este despliegue.** Es el "
            "archivo con el detalle fila a fila, y lleva ID_PREDIO y número "
            "predial de 288 mil predios, así que el `.gitignore` lo deja "
            "fuera del repositorio a propósito: este repositorio es público. "
            "Corriendo la app en local el explorador funciona completo; si no "
            "tiene el archivo, genérelo con `python src/comparacion_vigencia.py`.")
    else:
        det = cargar_detalle(str(RUTA_DETALLE), os.path.getmtime(RUTA_DETALLE))
        detf = filtrar(det)

        st.caption("Esto SÍ trae identificadores: ID_PREDIO y número predial. "
                   "Responde a los mismos filtros de la barra lateral que las "
                   "demás hojas, más los de aquí abajo.")

        f1, f2, f3 = st.columns([2, 1, 1])
        busca = f1.text_input(
            "Buscar por ID_PREDIO o número predial",
            placeholder="uno o varios, separados por espacio o coma",
            help="Coincidencia exacta en cualquiera de las dos columnas. Pasa "
                 "por encima de los filtros de la barra lateral para que un "
                 "predio concreto siempre aparezca.")
        rangos_disp = (sorted(detf["RANGO_VARIACION"].dropna().unique())
                       if "RANGO_VARIACION" in detf.columns else [])
        sel_rango = f2.multiselect("Rango de variación", rangos_disp)
        solo_fuera = f3.checkbox(
            "Solo fuera de tolerancia", value=False,
            help="La marca FUERA_TOLERANCIA de comparacion_vigencia.py.",
            disabled="FUERA_TOLERANCIA" not in detf.columns)

        if busca.strip():
            claves = {t.strip() for t in busca.replace(",", " ").split()
                      if t.strip()}
            detf = det[det["ID_PREDIO"].isin(claves)
                       | det["NUMERO_PREDIAL_NACIONAL"].isin(claves)]
            st.caption(f"Búsqueda directa sobre el detalle completo: "
                       f"{entero(len(detf))} filas para "
                       f"{entero(len(claves))} clave(s). Los filtros de la "
                       f"barra lateral no se aplican.")
        else:
            if sel_rango:
                detf = detf[detf["RANGO_VARIACION"].isin(sel_rango)]
            if solo_fuera and "FUERA_TOLERANCIA" in detf.columns:
                detf = detf[detf["FUERA_TOLERANCIA"].astype(str)
                            .isin(["1", "True", "true", "SI", "SÍ"])]

        if detf.empty:
            st.warning("Ninguna construcción cumple lo pedido.")
        else:
            # Columnas: las fijas mas las cuatro de la medida elegida arriba,
            # para que responda al radio de la barra lateral igual que el resto
            # y no haya que buscar entre cincuenta columnas.
            cols_medida = [c for c in (medida["vig"], medida["liq"],
                                       medida.get("dif"), medida["var"])
                           if c and c in detf.columns]
            cortas = ([c for c in COLUMNAS_DETALLE_FIJAS if c in detf.columns]
                      + cols_medida)

            c1, c2 = st.columns([1, 1])
            vista = c1.radio(
                "Columnas", ["Las de la medida elegida", "Todas"],
                horizontal=True,
                help=f"La vista corta deja los identificadores y las cuatro "
                     f"columnas de {etiqueta_medida.lower()} "
                     f"{etiqueta_base.lower()}. La completa trae las "
                     f"{len(detf.columns)} del archivo.")
            n_ver = c2.number_input(
                "Filas en pantalla", 50, 5000, 300, step=50,
                help="Solo cuántas se dibujan aquí. La descarga no depende "
                     "de esto.")

            cols = cortas if vista.startswith("Las") else list(detf.columns)
            st.caption(f"**{entero(len(detf))} construcciones** cumplen la "
                       f"selección · se muestran las primeras {entero(n_ver)}")
            st.dataframe(con_formato(detf[cols].head(int(n_ver))),
                         width="stretch", hide_index=True)

            st.divider()
            st.markdown("**Bajar esta selección**")

            # NADA se arma antes de que lo pidan. st.download_button construye
            # el dato ANTES de que nadie haga clic, y aqui eso serian 200 MB de
            # CSV -o tres minutos de Excel- en CADA interaccion con la pagina.
            # Por eso va en dos pasos, y lo preparado se guarda con la FIRMA de
            # la seleccion: si se cambia un filtro, el archivo listo deja de
            # ofrecerse en vez de entregar algo que ya no corresponde.
            g1, g2 = st.columns([1, 1])
            formato = g1.radio("Formato", ["Excel", "CSV"], horizontal=True,
                               help="El Excel va a unas 1.500 filas por "
                                    "segundo y trae la hoja Diccionario. El "
                                    "CSV no tiene tope y es mucho más rápido.")
            tope = g2.number_input(
                "Filas al Excel", 500, 50000, 5000, step=500,
                disabled=formato != "Excel",
                help="5.000 filas tardan ~4 s; 50.000, cerca de medio minuto.")

            recorte = detf.head(int(tope)) if formato == "Excel" else detf
            if formato == "Excel" and len(detf) > len(recorte):
                st.warning(f"El Excel llevará {entero(len(recorte))} de las "
                           f"{entero(len(detf))} filas. Para llevarlas todas, "
                           f"use el CSV o afine los filtros.")

            firma = (medida["prefijo"], tuple(sel_familia), tuple(sel_tabla),
                     tuple(sel_comuna), tuple(sel_actividad), tuple(sel_grupo),
                     busca.strip(), tuple(sel_rango), bool(solo_fuera),
                     formato, int(tope), len(detf))

            if st.button(f"🧾 Preparar el {formato}", type="secondary",
                         disabled=formato == "Excel" and not MOTOR_EXCEL,
                         help=None if MOTOR_EXCEL else
                         "Este entorno no tiene xlsxwriter instalado."):
                with st.spinner(f"Armando el {formato} "
                                f"({entero(len(recorte))} filas)…"):
                    st.session_state["descarga"] = {
                        "firma": firma,
                        "n": len(recorte),
                        "datos": (excel_predios(recorte) if formato == "Excel"
                                  else recorte.to_csv(index=False)
                                              .encode("utf-8-sig")),
                        "nombre": (f"DETALLE_PREDIOS_{V_BASE}_{V_LIQ}"
                                   f".{'xlsx' if formato == 'Excel' else 'csv'}"),
                        "mime": ("application/vnd.openxmlformats-officedocument"
                                 ".spreadsheetml.sheet" if formato == "Excel"
                                 else "text/csv")}

            listo = st.session_state.get("descarga")
            if listo and listo["firma"] == firma:
                st.download_button(
                    f"⬇️ Descargar {listo['nombre']} "
                    f"({entero(listo['n'])} filas · "
                    f"{_miles(len(listo['datos']) / 1e6, 1)} MB)",
                    data=listo["datos"], file_name=listo["nombre"],
                    mime=listo["mime"], type="primary", key="dl_predios")
            elif listo:
                st.caption("Cambió la selección: vuelva a preparar la "
                           "descarga para no bajar lo de antes.")



# ---------------------------------------------------------------------
# HOJA 3 - REGLAS
# ---------------------------------------------------------------------
with hoja_reglas:
    st.subheader("Cómo se asigna la tabla de valor")
    st.caption("Hoja informativa: es la regla **general**, no un recuento de "
               "los predios de esta corrida. No responde a los filtros de la "
               "izquierda. Se descarga con el botón de la hoja Tablas, en la "
               "pestaña REGLAS del Excel.")

    st.dataframe(reglas, width="stretch", hide_index=True,
                 column_config={"COMUNAS": st.column_config.TextColumn(
                     width="medium")})

    st.caption("Se entra por el uso de la construcción, se mira en qué grupo "
               "cae la comuna, luego la condición jurídica y la tipología de "
               "la ZHF: eso da la columna del Excel de tablas de valor. Dentro "
               "de esa columna, el VM2 se lee en la fila del puntaje "
               "(PUNTCONS), que va de 1 a 100.")

    st.divider()
    st.markdown("**Grupos de comunas**")
    st.markdown("\n".join(f"- **{g}** — {', '.join(c)}"
                          for g, c in GRUPOS_FILTRO.items()))
    st.caption("Las 5 comunas extra no tienen tabla propia: hoy se liquidan "
               "leyendo las mismas columnas *_10C_* del grupo de 10.")

    st.divider()
    st.markdown("**Excepciones**")
    st.info("**ZHF fuera de 011-016.** Si las tres últimas posiciones de la "
            "ZHF son diferentes a 011, 012, 013, 014, 015 y 016, se emplea el "
            "**estrato socioeconómico del predio (ESTRPRED)** y se asigna la "
            "tabla correspondiente según la comuna y la condición jurídica.")

    st.divider()
    st.markdown("**Qué predios NO se liquidan por tabla**")
    st.markdown(
        """
Tres razones, y ninguna entra a este reporte: comparar contra un VM2 de tabla
un valor que no salió de una tabla no mide la tabla.

**1. Van por modelo, no por tabla**

| Uso de construcción |
|---|
| `Apartamentos_4_y_mas_pisos_en_PH` |
| `Comercio_en_PH` |
| `Oficinas_Consultorios_en_PH` |

**2. Son especiales por defecto**

| Uso de construcción |
|---|
| `Centros_Comerciales_grandes` |
| `Centros_Comerciales_en_PH_grandes` |

**3. El predio tiene información incompleta**

Sin área, sin valor, sin puntaje o sin VM2 de tabla: no hay con qué comparar.
        """)




st.caption(
    f"Fuente: {RUTA_DATOS.name} · generado por comparacion_vigencia.py el "
    f"{pd.Timestamp(os.path.getmtime(RUTA_DATOS), unit='s'):%Y-%m-%d %H:%M}. "
    f"predial: el VM2 viene redondeado a $100 y el avalúo a $100.000, así que "
    f"puede haber diferencias de centésimas contra el reporte en Excel, que "
    f"trabaja con el valor exacto y es el documento de referencia."
)


# =====================================================================
# PUBLICAR (gratis, con enlace)
# =====================================================================
# Streamlit Community Cloud publica esto sin costo y da una URL fija que se
# abre desde cualquier PC:
#
#   1. En el repositorio de GitHub tienen que estar:
#          src/app_vigencias.py
#          output/COMPARACION_VIGENCIA_PUBLICO.parquet   (~7 MB)
#          requirements.txt
#      Y NO puede estar COMPARACION_VIGENCIA_DETALLE.parquet: lleva ID_PREDIO
#      y numero predial. El .gitignore ya lo deja fuera.
#   2. En share.streamlit.io, "Create app", elegir el repositorio y poner
#      src/app_vigencias.py como archivo principal.
#   3. Queda en https://<lo-que-se-elija>.streamlit.app
#
# OJO con el historial: borrar un archivo en un commit nuevo no lo saca de los
# commits anteriores, que se siguen pudiendo descargar. Si el detalle llego a
# subirse alguna vez, no basta con borrarlo: hay que rehacer el repositorio.
